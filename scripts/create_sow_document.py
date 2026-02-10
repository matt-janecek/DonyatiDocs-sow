#!/usr/bin/env python3
"""
Create a SOW Word document from a SOW pricing Excel workbook.

Usage:
    python create_sow_document.py pricing.xlsx output.docx [--scope scope.json]

This script reads a SOW pricing Excel workbook (created by donsow or manually)
and generates a professional Statement of Work Word document using the donword
template system.

The Excel workbook should have:
- Summary or Pricing Details sheet with project info
- Resource rows with roles, hours, and rates
- Optional: Timeline and Deliverables sheets

Optional scope.json can provide:
- Custom scope of services text
- Custom deliverables list
- Custom client responsibilities
- Custom assumptions
"""

import argparse
import json
import sys
import subprocess
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl


# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
TEMPLATES_DIR = PROJECT_ROOT / "templates" / "word"

# Default content sections
DEFAULT_ABOUT_SOW = """This Statement of Work ("SOW"), entered into as of {date}, is by and between Donyati, LLC ("Donyati") and {client_name} ("Client"). This SOW is governed by the terms and conditions of the Master Services Agreement ("MSA") dated {msa_date}, between Donyati and Client. In the event of any conflict between the terms of this SOW and the MSA, the terms of this SOW shall prevail with respect to the Services described herein.

The primary goal of this engagement is to provide {service_type} services for {project_name}. Donyati will leverage its expertise in {capability_area} to deliver the Services outlined in this SOW."""

DEFAULT_CLIENT_RESPONSIBILITIES = [
    "Provide a designated representative to act as the main point of Client contact for the Donyati Consultants.",
    "Confirm appropriate stakeholders and support resources are available as requested by Donyati consultants.",
    "Provide appropriate Application documentation access to Donyati resources.",
    "Provide timely responses to questions and clarifications within two (2) business days.",
    "All management functions and decisions related to Services, including without limitation evaluation and acceptance of Deliverables, will remain the responsibility of Client.",
    "Additional Client Responsibilities: In connection with Donyati's provision of Services, Client shall provide appropriate review and approval of deliverables within two (2) business days of receipt or as agreed to by both parties."
]

DEFAULT_ASSUMPTIONS = [
    "All services will be provided in English, including but not limited to project documents, presentations, communications, training, workshops, and meetings.",
    "All Donyati resources will work from Donyati sites (including home offices), VPN access and/or laptop will be provided as applicable.",
    "Standard business hours for US are 9:00 AM – 5:00 PM ET.",
    "Client will provide Donyati access to all necessary systems and environments to commence work in a timely manner.",
    "Donyati will not access Client Production source systems.",
    "In the event that the project start date, as specified in the SOW, is delayed by Client for more than thirty (30) days, Donyati reserves the right to re-estimate the project considering resource availability.",
    "If Client requests Donyati to begin work on this project prior to signatures, this SOW is deemed as signed by Client on the date work begins."
]

DEFAULT_OUT_OF_SCOPE = [
    "Services not explicitly included in this SOW are implicitly excluded from this engagement."
]

DEFAULT_MISC_PROVISIONS = [
    {"provision": "Purchase Order Number (PO#)", "narrative": "Client shall provide Donyati with a Purchase Order Number (PO#) prior to commencement of Services."},
    {"provision": "Scope Changes", "narrative": "Changes to the scope of services will be documented through a formal Change Order process and require written approval from both parties."},
    {"provision": "Deliverables", "narrative": "All deliverables provided to Client under this SOW shall be considered accepted unless Client provides written notice of rejection within five (5) business days of receipt."},
    {"provision": "Termination for Convenience", "narrative": "This SOW may be terminated for convenience by either party with thirty (30) days prior written notice."},
]


def read_excel_workbook(excel_path: str) -> Dict[str, Any]:
    """Read SOW pricing data from Excel workbook."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    data = {
        "project": {},
        "sales_team": {},
        "resources": [],
        "phases": [],
        "deliverables": [],
        "totals": {}
    }

    # Try to find Pricing Details or Summary sheet
    pricing_sheet = None
    for name in ["Pricing Details", "Summary", "Pricing Sheet"]:
        if name in wb.sheetnames:
            pricing_sheet = wb[name]
            break

    if not pricing_sheet:
        pricing_sheet = wb.active

    # Extract project info from Pricing Details sheet
    if "Pricing Details" in wb.sheetnames:
        ws = wb["Pricing Details"]

        # Row 1: Client Name, Pricing Date
        data["project"]["client_name"] = _get_cell_value(ws, "B1", "")
        data["project"]["pricing_date"] = _format_date(_get_cell_value(ws, "D1", ""))

        # Row 2: Capability, Sales Rep, Type
        data["project"]["capability_area"] = _get_cell_value(ws, "B2", "")
        data["sales_team"]["sales_rep"] = _get_cell_value(ws, "D2", "")
        data["project"]["project_type"] = _get_cell_value(ws, "F2", "New")

        # Row 3: Service Type, Presales/Delivery, Risk Profile
        data["project"]["service_type"] = _get_cell_value(ws, "B3", "")
        data["sales_team"]["relationship_owner"] = _get_cell_value(ws, "D3", "")
        data["project"]["risk_profile"] = _get_cell_value(ws, "F3", "Low")

        # Row 4: Contract Type, Engagement Lead, Project Start Date
        data["project"]["contract_type"] = _get_cell_value(ws, "B4", "T&M")
        data["sales_team"]["engagement_lead"] = _get_cell_value(ws, "D4", "")
        data["project"]["project_start_date"] = _format_date(_get_cell_value(ws, "F4", ""))

        # Extract resources (starting row 7)
        for row in range(7, ws.max_row + 1):
            practice = _get_cell_value(ws, f"A{row}", "")
            resource_role = _get_cell_value(ws, f"B{row}", "")

            if not practice and not resource_role:
                continue
            if practice == "" and resource_role == "":
                continue

            # Check for TOTAL row
            location = _get_cell_value(ws, f"E{row}", "")
            if location == "TOTAL" or practice == "TOTAL":
                break

            resource = {
                "practice": practice,
                "resource_role": resource_role,
                "project_role": _get_cell_value(ws, f"C{row}", ""),
                "potential_resource": _get_cell_value(ws, f"D{row}", "TBD"),
                "location": location,
                "hourly_rate": _get_cell_value(ws, f"F{row}", 0),
            }

            # Get monthly hours (columns G onwards)
            monthly_hours = []
            for col in range(7, 20):  # Up to 13 months
                hours = ws.cell(row=row, column=col).value
                if hours is not None and isinstance(hours, (int, float)):
                    monthly_hours.append(int(hours))
                elif isinstance(hours, str) and hours.strip():
                    try:
                        monthly_hours.append(int(float(hours)))
                    except:
                        break
                else:
                    break

            resource["monthly_hours"] = monthly_hours
            resource["total_hours"] = sum(monthly_hours)
            resource["total_fee"] = resource["total_hours"] * (resource["hourly_rate"] or 0)

            if resource["resource_role"]:  # Only add if has a role
                data["resources"].append(resource)

    # Try Summary sheet for additional info or totals
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        # Look for project name and totals
        for row in range(1, min(30, ws.max_row + 1)):
            label = _get_cell_value(ws, f"A{row}", "").lower()
            value = _get_cell_value(ws, f"B{row}", "")

            if "project name" in label:
                data["project"]["project_name"] = value
            elif "client" in label and "name" in label:
                if not data["project"].get("client_name"):
                    data["project"]["client_name"] = value
            elif "total hours" in label:
                data["totals"]["total_hours"] = value
            elif "total fee" in label:
                data["totals"]["total_fees"] = value

    # Extract deliverables if sheet exists
    if "Deliverables" in wb.sheetnames or "Deliverables " in wb.sheetnames:
        ws_name = "Deliverables" if "Deliverables" in wb.sheetnames else "Deliverables "
        ws = wb[ws_name]

        # Get phase headers from row 1
        phases = []
        for col in range(1, ws.max_column + 1):
            phase = _get_cell_value(ws, ws.cell(row=1, column=col).coordinate, "")
            if phase:
                phases.append({"name": phase.strip(), "deliverables": []})

        # Get deliverables for each phase
        for col_idx, phase in enumerate(phases):
            for row in range(3, ws.max_row + 1):
                item = ws.cell(row=row, column=col_idx + 1).value
                if item and str(item).strip():
                    phase["deliverables"].append(str(item).strip())

        data["deliverables"] = [p for p in phases if p["deliverables"]]

    # Extract timeline/phases if sheet exists
    if "Timeline 36 Weeks" in wb.sheetnames:
        ws = wb["Timeline 36 Weeks"]
        # Look for phase names in column A/B
        for row in range(6, min(25, ws.max_row + 1)):
            phase_name = ws.cell(row=row, column=1).value or ws.cell(row=row, column=2).value
            if phase_name and str(phase_name).strip():
                data["phases"].append({"name": str(phase_name).strip()})

    # Calculate totals if not already set
    if not data["totals"].get("total_hours"):
        data["totals"]["total_hours"] = sum(r.get("total_hours", 0) for r in data["resources"])
    if not data["totals"].get("total_fees"):
        data["totals"]["total_fees"] = sum(r.get("total_fee", 0) for r in data["resources"])

    # Infer project name if not found
    if not data["project"].get("project_name"):
        capability = data["project"].get("capability_area", "")
        service = data["project"].get("service_type", "")
        data["project"]["project_name"] = f"{capability} {service}".strip() or "Professional Services"

    return data


def _get_cell_value(ws, cell_ref: str, default=""):
    """Get cell value with default."""
    try:
        value = ws[cell_ref].value
        if value is None:
            return default
        return value
    except:
        return default


def _format_date(value) -> str:
    """Format date value to string."""
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%B %d, %Y")
    if isinstance(value, str):
        # Try to parse common formats
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
            try:
                dt = datetime.strptime(value, fmt)
                return dt.strftime("%B %d, %Y")
            except:
                continue
        return value
    return str(value)


def generate_sow_content(excel_data: Dict, scope_data: Optional[Dict] = None) -> Dict:
    """Generate donword-compatible content JSON from Excel data."""

    project = excel_data.get("project", {})
    sales_team = excel_data.get("sales_team", {})
    resources = excel_data.get("resources", [])
    deliverables = excel_data.get("deliverables", [])
    totals = excel_data.get("totals", {})

    client_name = project.get("client_name", "Client")
    project_name = project.get("project_name", "Professional Services")

    # Use scope data if provided, otherwise use defaults
    scope = scope_data or {}

    # Build the document content
    content = {
        "title": f"Statement of Work",
        "subtitle": project_name,
        "client": client_name,
        "date": project.get("pricing_date", datetime.now().strftime("%B %d, %Y")),
        "type": "proposal",
        "confidential": False,
        "header_text": f"SOW - {project_name}",
        "sections": []
    }

    # Header info section
    header_section = {
        "heading": "",
        "level": 1,
        "content": [
            {"type": "paragraph", "text": f"Client Name: {client_name}"},
            {"type": "paragraph", "text": f"Project Name: {project_name}"},
            {"type": "paragraph", "text": f"Date: {project.get('pricing_date', '')}"},
        ]
    }
    content["sections"].append(header_section)

    # About this Statement of Work
    about_text = scope.get("about", DEFAULT_ABOUT_SOW).format(
        date=project.get("pricing_date", "the date hereof"),
        client_name=client_name,
        msa_date=scope.get("msa_date", "[MSA DATE]"),
        service_type=project.get("service_type", "professional"),
        project_name=project_name,
        capability_area=project.get("capability_area", "technology consulting")
    )

    content["sections"].append({
        "heading": "About this Statement of Work",
        "level": 1,
        "content": [{"type": "paragraph", "text": about_text}]
    })

    # Scope of Services
    scope_content = []
    scope_intro = scope.get("scope_intro", f"Under this SOW, Donyati will work with Client to provide Services noted below.")
    scope_content.append({"type": "paragraph", "text": scope_intro})

    if scope.get("scope_items"):
        scope_content.append({"type": "bullets", "items": scope.get("scope_items")})
    elif resources:
        # Generate scope from resources
        scope_items = []
        roles_seen = set()
        for r in resources:
            role = r.get("project_role") or r.get("resource_role", "")
            if role and role not in roles_seen:
                scope_items.append(f"Provide {role} services as outlined in this SOW")
                roles_seen.add(role)
        if scope_items:
            scope_content.append({"type": "bullets", "items": scope_items})

    content["sections"].append({
        "heading": "Scope of Services",
        "level": 1,
        "content": scope_content
    })

    # Deliverables
    deliverables_content = []
    if scope.get("deliverables"):
        deliverables_content.append({"type": "bullets", "items": scope.get("deliverables")})
    elif deliverables:
        for phase in deliverables:
            if phase.get("deliverables"):
                deliverables_content.append({"type": "heading", "text": phase["name"], "level": 2})
                deliverables_content.append({"type": "bullets", "items": phase["deliverables"]})
    else:
        deliverables_content.append({"type": "paragraph", "text": "Deliverables will be defined during the engagement based on agreed scope."})

    content["sections"].append({
        "heading": "Deliverables",
        "level": 1,
        "content": deliverables_content
    })

    # Client Responsibilities
    responsibilities = scope.get("client_responsibilities", DEFAULT_CLIENT_RESPONSIBILITIES)
    content["sections"].append({
        "heading": "Client Responsibilities",
        "level": 1,
        "content": [
            {"type": "paragraph", "text": "Donyati will work directly with the Client to perform Services under this SOW. To ensure the project is successful, Client agrees to:"},
            {"type": "bullets", "items": responsibilities}
        ]
    })

    # Engagement Fees
    fees_content = []

    contract_type = project.get("contract_type", "T&M")
    start_date = project.get("project_start_date", "[START DATE]")

    fees_intro = f"Donyati shall charge, and Client shall pay for the Services under this SOW based on {contract_type} fees."
    fees_content.append({"type": "paragraph", "text": fees_intro})
    fees_content.append({"type": "paragraph", "text": f"The Service Commencement Date is {start_date}."})

    # Resource table
    if resources:
        table_headers = ["Resource Role", "Location", "Hours", "Hourly Rate", "Estimated Fee"]
        table_rows = []

        for r in resources:
            role = r.get("project_role") or r.get("resource_role", "")
            location = r.get("location", "USA")
            hours = r.get("total_hours", 0)
            rate = r.get("hourly_rate", 0)
            fee = hours * rate

            table_rows.append([
                role,
                location,
                f"{hours:,}",
                f"${rate:,}",
                f"${fee:,}"
            ])

        # Add total row
        total_hours = totals.get("total_hours", sum(r.get("total_hours", 0) for r in resources))
        total_fees = totals.get("total_fees", sum(r.get("total_hours", 0) * r.get("hourly_rate", 0) for r in resources))
        table_rows.append(["TOTAL", "", f"{total_hours:,}", "", f"${total_fees:,}"])

        fees_content.append({"type": "table", "headers": table_headers, "rows": table_rows})

    fees_content.append({
        "type": "paragraph",
        "text": "Invoices will be provided monthly at the end of the service month, and payment is due in U.S. currency within thirty (30) days of receipt."
    })

    content["sections"].append({
        "heading": "Engagement Fees",
        "level": 1,
        "content": fees_content
    })

    # Assumptions
    assumptions = scope.get("assumptions", DEFAULT_ASSUMPTIONS)
    content["sections"].append({
        "heading": "Assumptions",
        "level": 1,
        "content": [
            {"type": "paragraph", "text": "The following general assumptions have been made in the development of this SOW:"},
            {"type": "bullets", "items": assumptions}
        ]
    })

    # Outside Scope
    out_of_scope = scope.get("out_of_scope", DEFAULT_OUT_OF_SCOPE)
    content["sections"].append({
        "heading": "Outside the Scope of this SOW",
        "level": 1,
        "content": [{"type": "bullets", "items": out_of_scope}]
    })

    # Miscellaneous Provisions
    misc_provisions = scope.get("misc_provisions", DEFAULT_MISC_PROVISIONS)
    if misc_provisions:
        misc_content = []
        misc_table_rows = [[p["provision"], p["narrative"]] for p in misc_provisions]
        misc_content.append({
            "type": "table",
            "headers": ["Provision", "Narrative"],
            "rows": misc_table_rows
        })

        content["sections"].append({
            "heading": "Miscellaneous Provisions",
            "level": 1,
            "content": misc_content
        })

    return content


def create_sow_document(excel_path: str, output_path: str, scope_path: Optional[str] = None,
                        use_donword: bool = True) -> str:
    """Create SOW Word document from Excel pricing workbook."""

    # Read Excel data
    print(f"Reading Excel workbook: {excel_path}")
    excel_data = read_excel_workbook(excel_path)

    print(f"  Client: {excel_data['project'].get('client_name', 'N/A')}")
    print(f"  Project: {excel_data['project'].get('project_name', 'N/A')}")
    print(f"  Resources: {len(excel_data['resources'])}")

    # Load scope data if provided
    scope_data = None
    if scope_path and Path(scope_path).exists():
        print(f"Loading scope data: {scope_path}")
        with open(scope_path, 'r', encoding='utf-8') as f:
            scope_data = json.load(f)

    # Generate content JSON
    content = generate_sow_content(excel_data, scope_data)

    if use_donword:
        # Write temp JSON and call create_document.py
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
            json.dump(content, f, indent=2)
            temp_json = f.name

        try:
            print(f"Generating Word document using donword...")
            create_doc_script = SCRIPT_DIR / "create_document.py"

            result = subprocess.run(
                [sys.executable, str(create_doc_script), temp_json, output_path, "--template", "cover"],
                capture_output=True,
                text=True
            )

            if result.returncode != 0:
                print(f"Warning: create_document.py returned non-zero: {result.stderr}")

            print(result.stdout)

        finally:
            Path(temp_json).unlink(missing_ok=True)
    else:
        # Just output the JSON for inspection
        output_json = Path(output_path).with_suffix('.json')
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(content, f, indent=2)
        print(f"Content JSON written to: {output_json}")
        return str(output_json)

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Create SOW Word document from pricing Excel workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python create_sow_document.py pricing.xlsx output.docx
  python create_sow_document.py pricing.xlsx output.docx --scope scope.json
  python create_sow_document.py pricing.xlsx output.json --json-only
        """
    )
    parser.add_argument("excel", help="Path to SOW pricing Excel workbook")
    parser.add_argument("output", help="Path for output DOCX (or JSON with --json-only)")
    parser.add_argument("--scope", help="Optional JSON file with scope/deliverables content")
    parser.add_argument("--json-only", action="store_true", help="Output content JSON instead of Word doc")

    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"Error: Excel file not found: {args.excel}")
        sys.exit(1)

    try:
        output = create_sow_document(
            str(excel_path),
            args.output,
            args.scope,
            use_donword=not args.json_only
        )
        print(f"\nCreated: {output}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
