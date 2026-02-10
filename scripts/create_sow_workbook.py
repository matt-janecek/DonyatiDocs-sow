#!/usr/bin/env python3
"""
Create a SOW pricing Excel workbook from JSON specification.

Usage:
    # From JSON input
    python create_sow_workbook.py sow-data.json output.xlsx

    # Clone and modify existing workbook
    python create_sow_workbook.py --clone existing.xlsx output.xlsx --json modifications.json

The sow-data.json should have the structure:
{
  "project": {
    "client_name": "IEEE",
    "project_name": "AI Chatbot Implementation",
    "capability_area": "Cloud Engineering",
    "service_type": "Implementation",
    "contract_type": "T&M",
    "project_type": "New",
    "risk_profile": "Low",
    "pricing_date": "2026-01-29",
    "project_start_date": "2026-02-15",
    "duration_months": 6
  },
  "sales_team": {
    "relationship_owner": "Name",
    "sales_rep": "Name",
    "inside_sales": "Name",
    "sales_team_leader": "Name"
  },
  "resources": [
    {
      "practice": "EPM_Practice_USA",
      "resource_role": "Project Manager",
      "project_role": "AI Scrum Master",
      "potential_resource": "TBD",
      "location": "USA",
      "hourly_rate": 230,
      "monthly_hours": [24, 168, 160, 176, 168, 168]
    }
  ],
  "phases": [
    {"name": "Mobilize", "start_week": 1, "end_week": 2},
    {"name": "Requirements", "start_week": 2, "end_week": 5}
  ],
  "deliverables": "default"
}
"""

import argparse
import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from copy import copy

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
DATA_DIR = PROJECT_ROOT / "data"
TEMPLATES_DIR = PROJECT_ROOT / "templates" / "excel"
REFERENCE_DATA_PATH = DATA_DIR / "sow-reference-data.json"

# Donyati brand colors
DONYATI_PURPLE = "4A4778"
DONYATI_BLACK = "12002A"
DONYATI_LIGHT_PURPLE = "E8E6F0"
HEADER_BG = "4A4778"
ALT_ROW_BG = "F5F0FA"


def load_reference_data() -> Dict[str, Any]:
    """Load picklist/reference data from JSON file."""
    if not REFERENCE_DATA_PATH.exists():
        print(f"Warning: Reference data not found at {REFERENCE_DATA_PATH}")
        return {}

    with open(REFERENCE_DATA_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_rate(ref_data: Dict, practice: str, role: str) -> int:
    """Look up billing rate for a practice/role combination."""
    key = f"{practice}|{role}"
    if "rates" in ref_data and key in ref_data["rates"]:
        return ref_data["rates"][key]["bill_rate"]
    # Default rates by location
    if "India" in practice:
        return 55
    return 200


def create_styles() -> Dict[str, NamedStyle]:
    """Create named styles for the workbook."""
    styles = {}

    # Header style
    styles["header"] = NamedStyle(name="sow_header")
    styles["header"].font = Font(bold=True, color="FFFFFF", size=10)
    styles["header"].fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    styles["header"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    styles["header"].border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Data cell style
    styles["data"] = NamedStyle(name="sow_data")
    styles["data"].font = Font(size=10)
    styles["data"].alignment = Alignment(horizontal="left", vertical="center")
    styles["data"].border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Number style
    styles["number"] = NamedStyle(name="sow_number")
    styles["number"].font = Font(size=10)
    styles["number"].alignment = Alignment(horizontal="right", vertical="center")
    styles["number"].border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    styles["number"].number_format = "#,##0"

    # Currency style
    styles["currency"] = NamedStyle(name="sow_currency")
    styles["currency"].font = Font(size=10)
    styles["currency"].alignment = Alignment(horizontal="right", vertical="center")
    styles["currency"].border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    styles["currency"].number_format = "$#,##0"

    # Label style (purple background for editable cells)
    styles["label"] = NamedStyle(name="sow_label")
    styles["label"].font = Font(bold=True, size=10)
    styles["label"].fill = PatternFill(start_color=DONYATI_LIGHT_PURPLE, end_color=DONYATI_LIGHT_PURPLE, fill_type="solid")
    styles["label"].alignment = Alignment(horizontal="left", vertical="center")

    return styles


def apply_header_style(cell, styles: Dict):
    """Apply header styling to a cell."""
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


def apply_data_style(cell, styles: Dict, is_alt_row: bool = False):
    """Apply data cell styling."""
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    if is_alt_row:
        cell.fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")


def create_summary_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Summary sheet with project overview and totals."""
    ws.title = "Summary"

    project = data.get("project", {})
    sales = data.get("sales_team", {})
    resources = data.get("resources", [])

    # Title
    ws["A1"] = "SOW Pricing Summary"
    ws["A1"].font = Font(bold=True, size=16, color=DONYATI_BLACK)
    ws.merge_cells("A1:D1")

    # General Information section
    ws["A3"] = "General Information"
    ws["A3"].font = Font(bold=True, size=12, color=DONYATI_PURPLE)

    labels = [
        ("A4", "Client Name:", "B4", project.get("client_name", "")),
        ("A5", "Project Name:", "B5", project.get("project_name", "")),
        ("A6", "Capability:", "B6", project.get("capability_area", "")),
        ("A7", "Service:", "B7", project.get("service_type", "")),
        ("A8", "Contract Type:", "B8", project.get("contract_type", "")),
        ("A9", "Risk Profile:", "B9", project.get("risk_profile", "")),
        ("A10", "Pricing Date:", "B10", project.get("pricing_date", "")),
        ("A11", "Project Start:", "B11", project.get("project_start_date", "")),
        ("A12", "Duration:", "B12", f"{project.get('duration_months', 0)} months"),
    ]

    for label_cell, label, value_cell, value in labels:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, size=10)
        ws[value_cell] = value

    # Sales Team section
    ws["A14"] = "Sales Team"
    ws["A14"].font = Font(bold=True, size=12, color=DONYATI_PURPLE)

    sales_labels = [
        ("A15", "Relationship Owner:", "B15", sales.get("relationship_owner", "")),
        ("A16", "Sales Rep:", "B16", sales.get("sales_rep", "")),
        ("A17", "Inside Sales:", "B17", sales.get("inside_sales", "")),
        ("A18", "Sales Team Leader:", "B18", sales.get("sales_team_leader", "")),
    ]

    for label_cell, label, value_cell, value in sales_labels:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True, size=10)
        ws[value_cell] = value

    # Resource Summary section
    ws["A20"] = "Resource Summary"
    ws["A20"].font = Font(bold=True, size=12, color=DONYATI_PURPLE)

    # Calculate totals
    total_hours = 0
    total_fees = 0
    for resource in resources:
        hours = sum(resource.get("monthly_hours", []))
        rate = resource.get("hourly_rate", get_rate(ref_data, resource.get("practice", ""), resource.get("resource_role", "")))
        total_hours += hours
        total_fees += hours * rate

    ws["A21"] = "Total Resources:"
    ws["B21"] = len(resources)
    ws["A22"] = "Total Hours:"
    ws["B22"] = total_hours
    ws["B22"].number_format = "#,##0"
    ws["A23"] = "Total Fees:"
    ws["B23"] = total_fees
    ws["B23"].number_format = "$#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def create_pricing_details_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Pricing Details sheet with resources and hours."""
    ws.title = "Pricing Details"

    project = data.get("project", {})
    sales = data.get("sales_team", {})
    resources = data.get("resources", [])
    duration_months = project.get("duration_months", 6)

    # Project metadata header
    ws["A1"] = "Client Name (Partner Name):"
    ws["B1"] = project.get("client_name", "")
    ws["C1"] = "Pricing Date:"
    ws["D1"] = project.get("pricing_date", "")

    ws["A2"] = "Capability/Practice Area:"
    ws["B2"] = project.get("capability_area", "")
    ws["C2"] = "Sales Rep:"
    ws["D2"] = sales.get("sales_rep", "")
    ws["E2"] = "Type:"
    ws["F2"] = project.get("project_type", "New")

    ws["A3"] = "Service /Project Type:"
    ws["B3"] = project.get("service_type", "")
    ws["C3"] = "Presales/Delivery:"
    ws["D3"] = sales.get("relationship_owner", "")
    ws["E3"] = "Risk Profile:"
    ws["F3"] = project.get("risk_profile", "Low")

    ws["A4"] = "Contract Type:"
    ws["B4"] = project.get("contract_type", "T&M")
    ws["C4"] = "Engagement Lead:"
    ws["D4"] = sales.get("sales_team_leader", "")
    ws["E4"] = "Project Start Date:"
    ws["F4"] = project.get("project_start_date", "")

    # Style the metadata section
    for row in range(1, 5):
        for col in ["A", "C", "E"]:
            cell = ws[f"{col}{row}"]
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color=DONYATI_LIGHT_PURPLE, end_color=DONYATI_LIGHT_PURPLE, fill_type="solid")

    # Resource table header row
    header_row = 6
    headers = ["Practice", "Resource Role", "Project Role", "Potential Resource", "Location", "Hourly Rate"]

    # Add month columns
    for month in range(1, duration_months + 1):
        headers.append(f"Month {month}")

    headers.extend(["Total Hours", "Total Fees"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        apply_header_style(cell, styles)

    # Resource rows
    for row_idx, resource in enumerate(resources, header_row + 1):
        practice = resource.get("practice", "")
        resource_role = resource.get("resource_role", "")
        project_role = resource.get("project_role", "")
        potential_resource = resource.get("potential_resource", "TBD")
        location = resource.get("location", "USA")
        hourly_rate = resource.get("hourly_rate", get_rate(ref_data, practice, resource_role))
        monthly_hours = resource.get("monthly_hours", [0] * duration_months)

        is_alt = (row_idx - header_row) % 2 == 0

        ws.cell(row=row_idx, column=1, value=practice)
        apply_data_style(ws.cell(row=row_idx, column=1), styles, is_alt)

        ws.cell(row=row_idx, column=2, value=resource_role)
        apply_data_style(ws.cell(row=row_idx, column=2), styles, is_alt)

        ws.cell(row=row_idx, column=3, value=project_role)
        apply_data_style(ws.cell(row=row_idx, column=3), styles, is_alt)

        ws.cell(row=row_idx, column=4, value=potential_resource)
        apply_data_style(ws.cell(row=row_idx, column=4), styles, is_alt)

        ws.cell(row=row_idx, column=5, value=location)
        apply_data_style(ws.cell(row=row_idx, column=5), styles, is_alt)

        rate_cell = ws.cell(row=row_idx, column=6, value=hourly_rate)
        rate_cell.number_format = "$#,##0"
        apply_data_style(rate_cell, styles, is_alt)

        # Monthly hours
        for month_idx, hours in enumerate(monthly_hours):
            col = 7 + month_idx
            hours_cell = ws.cell(row=row_idx, column=col, value=hours if hours else "")
            hours_cell.number_format = "#,##0"
            apply_data_style(hours_cell, styles, is_alt)

        # Fill remaining month columns with empty
        for month_idx in range(len(monthly_hours), duration_months):
            col = 7 + month_idx
            apply_data_style(ws.cell(row=row_idx, column=col), styles, is_alt)

        # Total Hours formula
        first_hour_col = get_column_letter(7)
        last_hour_col = get_column_letter(6 + duration_months)
        total_hours_col = 7 + duration_months
        hours_formula = f"=SUM({first_hour_col}{row_idx}:{last_hour_col}{row_idx})"
        total_hours_cell = ws.cell(row=row_idx, column=total_hours_col, value=hours_formula)
        total_hours_cell.number_format = "#,##0"
        apply_data_style(total_hours_cell, styles, is_alt)

        # Total Fees formula
        total_fees_col = total_hours_col + 1
        rate_col = get_column_letter(6)
        total_hours_col_letter = get_column_letter(total_hours_col)
        fees_formula = f"={rate_col}{row_idx}*{total_hours_col_letter}{row_idx}"
        total_fees_cell = ws.cell(row=row_idx, column=total_fees_col, value=fees_formula)
        total_fees_cell.number_format = "$#,##0"
        apply_data_style(total_fees_cell, styles, is_alt)

    # Totals row
    if resources:
        totals_row = header_row + len(resources) + 1
        ws.cell(row=totals_row, column=5, value="TOTAL")
        ws.cell(row=totals_row, column=5).font = Font(bold=True, size=10)

        # Total hours
        total_hours_col = 7 + duration_months
        first_data_row = header_row + 1
        last_data_row = header_row + len(resources)
        total_hours_col_letter = get_column_letter(total_hours_col)
        hours_sum = f"=SUM({total_hours_col_letter}{first_data_row}:{total_hours_col_letter}{last_data_row})"
        total_cell = ws.cell(row=totals_row, column=total_hours_col, value=hours_sum)
        total_cell.number_format = "#,##0"
        total_cell.font = Font(bold=True)

        # Total fees
        total_fees_col = total_hours_col + 1
        total_fees_col_letter = get_column_letter(total_fees_col)
        fees_sum = f"=SUM({total_fees_col_letter}{first_data_row}:{total_fees_col_letter}{last_data_row})"
        fees_cell = ws.cell(row=totals_row, column=total_fees_col, value=fees_sum)
        fees_cell.number_format = "$#,##0"
        fees_cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    for col in range(7, 7 + duration_months + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_timeline_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Timeline sheet with Gantt-style phase display."""
    ws.title = "Timeline 36 Weeks"

    project = data.get("project", {})
    phases = data.get("phases", [])
    duration_months = project.get("duration_months", 6)
    duration_weeks = duration_months * 4 + 4  # Extra buffer

    # Default phases if none provided
    if not phases:
        phases = [
            {"name": "Mobilize", "start_week": 1, "end_week": 2},
            {"name": "Requirements", "start_week": 2, "end_week": 5},
            {"name": "Design", "start_week": 4, "end_week": 8},
            {"name": "Build", "start_week": 7, "end_week": 16},
            {"name": "Testing", "start_week": 14, "end_week": 20},
            {"name": "Training", "start_week": 18, "end_week": 22},
            {"name": "Parallel Testing", "start_week": 20, "end_week": 24},
            {"name": "HyperCare", "start_week": 24, "end_week": 28},
        ]

    # Title
    ws["A1"] = "Project Timeline"
    ws["A1"].font = Font(bold=True, size=14, color=DONYATI_BLACK)

    # Week header row
    ws["A4"] = "Week #"
    ws["A4"].font = Font(bold=True, size=10)

    for week in range(1, min(37, duration_weeks + 1)):
        col = week + 1
        cell = ws.cell(row=4, column=col, value=week)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 4

    # Phase header
    ws["A5"] = "Project Phases"
    ws["A5"].font = Font(bold=True, size=10)

    # Phase rows
    phase_colors = [
        "4A4778",  # Purple
        "00629B",  # Blue
        "007377",  # Teal
        "009CA6",  # Cyan
        "FFA300",  # Amber
        "861F41",  # Maroon
        "2E7D32",  # Green
        "1976D2",  # Light Blue
    ]

    for row_idx, phase in enumerate(phases, 6):
        ws.cell(row=row_idx, column=1, value=phase["name"])
        ws.cell(row=row_idx, column=1).font = Font(size=10)

        start_week = phase.get("start_week", 1)
        end_week = phase.get("end_week", start_week + 4)
        color = phase_colors[(row_idx - 6) % len(phase_colors)]

        for week in range(start_week, min(end_week + 1, 37)):
            col = week + 1
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws.column_dimensions["A"].width = 20


def create_deliverables_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Deliverables sheet with phase-based deliverable matrix."""
    ws.title = "Deliverables"

    deliverables_setting = data.get("deliverables", "default")

    if deliverables_setting == "default" and "deliverables_by_phase" in ref_data:
        deliverables = ref_data["deliverables_by_phase"]
    elif isinstance(deliverables_setting, dict):
        deliverables = deliverables_setting
    else:
        deliverables = {
            "Mobilize": ["Project Charter", "RAID Logs", "Project Schedule"],
            "Requirements": ["RTM", "Requirements Workshops"],
            "Design": ["Design Document", "Architecture Diagram"],
            "Build": ["Application Development", "Data Integrations"],
            "Testing": ["Test Plan", "UAT Scripts"],
            "Training": ["Training Materials"],
            "Parallel Testing": ["Parallel Test Support"],
            "HyperCare": ["Production Support"],
        }

    # Phase headers
    phases = list(deliverables.keys())
    for col_idx, phase in enumerate(phases, 1):
        cell = ws.cell(row=1, column=col_idx, value=phase)
        apply_header_style(cell, styles)
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    # Find max deliverables
    max_items = max(len(items) for items in deliverables.values())

    # Deliverable items
    for col_idx, phase in enumerate(phases, 1):
        items = deliverables.get(phase, [])
        for row_idx, item in enumerate(items, 3):
            cell = ws.cell(row=row_idx, column=col_idx, value=item)
            apply_data_style(cell, styles, row_idx % 2 == 0)


def create_complexity_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Complexity Considerations sheet."""
    ws.title = "Complexity Considerations"

    ws["A1"] = "Complexity Considerations"
    ws["A1"].font = Font(bold=True, size=14, color=DONYATI_BLACK)

    considerations = [
        ("Data Volume", "Number of data sources, records, and integrations"),
        ("Integration Complexity", "Number and complexity of system integrations"),
        ("Business Process Complexity", "Complexity of workflows and business rules"),
        ("Organizational Change", "Level of organizational change management required"),
        ("Technical Environment", "Complexity of technical infrastructure"),
        ("Regulatory Requirements", "Compliance and regulatory considerations"),
        ("Timeline Constraints", "Aggressive timeline or fixed deadlines"),
        ("Resource Availability", "Client resource availability for UAT, training"),
        ("Geographic Distribution", "Multi-location or global deployment"),
        ("Customization Level", "Amount of custom development required"),
    ]

    ws["A3"] = "Factor"
    ws["B3"] = "Description"
    apply_header_style(ws["A3"], styles)
    apply_header_style(ws["B3"], styles)

    for row_idx, (factor, description) in enumerate(considerations, 4):
        ws.cell(row=row_idx, column=1, value=factor)
        ws.cell(row=row_idx, column=2, value=description)
        apply_data_style(ws.cell(row=row_idx, column=1), styles, row_idx % 2 == 0)
        apply_data_style(ws.cell(row=row_idx, column=2), styles, row_idx % 2 == 0)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60


def create_picklist_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Picklist sheet with reference data."""
    ws.title = "Picklist"

    # Headers
    headers = ["Client", "Capability", "Service Type", "Contract Type", "Practice", "Role"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell, styles)

    # Data columns
    lists = {
        1: ref_data.get("clients", []),
        2: ref_data.get("capabilities", []),
        3: ref_data.get("service_types", []),
        4: ref_data.get("contract_types", []),
        5: ref_data.get("practices", []),
        6: ref_data.get("roles", []),
    }

    max_rows = max(len(lst) for lst in lists.values()) if lists.values() else 0

    for col_idx, items in lists.items():
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=col_idx, value=item)

    # Column widths
    widths = [30, 25, 25, 15, 35, 30]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_instructions_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Instructions sheet."""
    ws.title = "INSTRUCTIONS"

    ws["A1"] = "SOW Pricing Workbook Instructions"
    ws["A1"].font = Font(bold=True, size=16, color=DONYATI_BLACK)

    instructions = [
        "",
        "Overview",
        "This workbook is used to develop pricing for Statement of Work (SOW) engagements.",
        "",
        "Key Sheets:",
        "- Summary: High-level project overview and totals",
        "- Pricing Details: Resource allocation and monthly hours",
        "- Timeline: Gantt-style project phase timeline",
        "- Deliverables: Phase-by-phase deliverable matrix",
        "- Complexity Considerations: Factors affecting estimate",
        "- Picklist: Reference data for dropdowns",
        "",
        "How to Use:",
        "1. Fill in project information in Pricing Details (rows 1-4)",
        "2. Add resources starting at row 7",
        "3. Enter monthly hours for each resource",
        "4. Review Timeline and adjust phases as needed",
        "5. Verify Deliverables match project scope",
        "6. Check Summary sheet for totals",
        "",
        "Notes:",
        "- Hourly rates are looked up from Picklist based on Practice + Role",
        "- Total Hours and Total Fees are calculated automatically",
        "- Purple-highlighted cells indicate editable fields",
    ]

    for row_idx, line in enumerate(instructions, 3):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if line and not line.startswith("-") and not line.startswith(" "):
            if line in ["Overview", "Key Sheets:", "How to Use:", "Notes:"]:
                cell.font = Font(bold=True, size=12, color=DONYATI_PURPLE)
            else:
                cell.font = Font(size=10)

    ws.column_dimensions["A"].width = 80


def create_resource_instructions_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Update Resource Instructions sheet."""
    ws.title = "Update Resource Instructions"

    ws["A1"] = "How to Update Resource List"
    ws["A1"].font = Font(bold=True, size=14, color=DONYATI_BLACK)

    instructions = [
        "",
        "The Certinia Resource List sheet contains available resources.",
        "",
        "To update:",
        "1. Export resource list from Certinia",
        "2. Copy data to Certinia Resource List sheet",
        "3. Ensure columns match: Account, Practice, First Name, Last Name, Role, Type",
        "",
        "The Picklist sheet will reference this data for dropdowns.",
    ]

    for row_idx, line in enumerate(instructions, 3):
        ws.cell(row=row_idx, column=1, value=line)

    ws.column_dimensions["A"].width = 70


def create_certinia_resource_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Certinia Resource List sheet."""
    ws.title = "Certinia Resource List"

    headers = ["Account Name", "Practice", "First Name", "Last Name", "Resource Role", "Employment Type"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        apply_header_style(cell, styles)
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    # Placeholder row
    ws.cell(row=2, column=1, value="Donyati")
    ws.cell(row=2, column=2, value="(Import from Certinia)")


def create_capability_service_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Customer Capability Service sheet."""
    ws.title = "Customer Capability Service"

    ws["A1"] = "Customer Capability Service Matrix"
    ws["A1"].font = Font(bold=True, size=14, color=DONYATI_BLACK)

    # Add capability/service matrix
    capabilities = ref_data.get("capabilities", [])[:20]
    services = ref_data.get("service_types", [])[:15]

    # Headers
    ws.cell(row=3, column=1, value="Capability \\ Service")
    apply_header_style(ws.cell(row=3, column=1), styles)

    for col_idx, service in enumerate(services, 2):
        cell = ws.cell(row=3, column=col_idx, value=service)
        apply_header_style(cell, styles)
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    for row_idx, capability in enumerate(capabilities, 4):
        ws.cell(row=row_idx, column=1, value=capability)
        apply_data_style(ws.cell(row=row_idx, column=1), styles, row_idx % 2 == 0)

    ws.column_dimensions["A"].width = 25


def create_formulas_sheet(ws, data: Dict, ref_data: Dict, styles: Dict):
    """Create the Formulas helper sheet."""
    ws.title = "Formulas"

    ws["A1"] = "Formula Reference Sheet"
    ws["A1"].font = Font(bold=True, size=14, color=DONYATI_BLACK)

    formulas = [
        ("Total Hours", "=SUM(range)", "Sum of all monthly hours for a resource"),
        ("Total Fees", "=Rate * Hours", "Hourly rate multiplied by total hours"),
        ("Blended Rate", "=Total Fees / Total Hours", "Average rate across all resources"),
        ("Margin %", "=(Bill Rate - Cost) / Bill Rate", "Profit margin percentage"),
    ]

    ws["A3"] = "Formula"
    ws["B3"] = "Syntax"
    ws["C3"] = "Description"
    apply_header_style(ws["A3"], styles)
    apply_header_style(ws["B3"], styles)
    apply_header_style(ws["C3"], styles)

    for row_idx, (name, syntax, desc) in enumerate(formulas, 4):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=syntax)
        ws.cell(row=row_idx, column=3, value=desc)
        for col in range(1, 4):
            apply_data_style(ws.cell(row=row_idx, column=col), styles, row_idx % 2 == 0)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45


def generate_workbook(input_path: str, template_path: Optional[str] = None) -> openpyxl.Workbook:
    """Generate a SOW workbook from JSON input."""
    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    ref_data = load_reference_data()
    styles = create_styles()

    if template_path and Path(template_path).exists():
        wb = openpyxl.load_workbook(template_path)
        # Clear existing sheets except template structure
        for sheet_name in wb.sheetnames:
            if sheet_name not in ["Formulas", "Picklist", "INSTRUCTIONS"]:
                del wb[sheet_name]
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Create all sheets
    sheets_config = [
        ("Summary", create_summary_sheet),
        ("Pricing Details", create_pricing_details_sheet),
        ("Timeline 36 Weeks", create_timeline_sheet),
        ("Deliverables", create_deliverables_sheet),
        ("Complexity Considerations", create_complexity_sheet),
        ("Formulas", create_formulas_sheet),
        ("Picklist", create_picklist_sheet),
        ("INSTRUCTIONS", create_instructions_sheet),
        ("Update Resource Instructions", create_resource_instructions_sheet),
        ("Certinia Resource List", create_certinia_resource_sheet),
        ("Customer Capability Service", create_capability_service_sheet),
    ]

    for sheet_name, create_func in sheets_config:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        create_func(ws, data, ref_data, styles)

    # Reorder sheets
    desired_order = [name for name, _ in sheets_config]
    for idx, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=idx - wb.sheetnames.index(sheet_name))

    return wb


def clone_workbook(source_path: str, modifications_path: Optional[str] = None) -> openpyxl.Workbook:
    """Clone an existing workbook and apply modifications."""
    wb = openpyxl.load_workbook(source_path)

    if modifications_path and Path(modifications_path).exists():
        with open(modifications_path, 'r', encoding='utf-8') as f:
            mods = json.load(f)

        # Apply modifications
        if "Pricing Details" in wb.sheetnames:
            ws = wb["Pricing Details"]
            project = mods.get("project", {})

            # Update project fields
            if "client_name" in project:
                ws["B1"] = project["client_name"]
            if "pricing_date" in project:
                ws["D1"] = project["pricing_date"]
            if "capability_area" in project:
                ws["B2"] = project["capability_area"]
            if "service_type" in project:
                ws["B3"] = project["service_type"]
            if "project_start_date" in project:
                ws["F4"] = project["project_start_date"]

    return wb


def main():
    parser = argparse.ArgumentParser(
        description="Create SOW pricing Excel workbook from JSON specification.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python create_sow_workbook.py sow-data.json output.xlsx
  python create_sow_workbook.py --clone existing.xlsx output.xlsx
  python create_sow_workbook.py --clone existing.xlsx output.xlsx --json modifications.json
        """
    )
    parser.add_argument("input", nargs="?", help="Path to SOW data JSON file")
    parser.add_argument("output", help="Path for output XLSX file")
    parser.add_argument("--clone", help="Clone mode: source workbook to copy")
    parser.add_argument("--json", help="Clone mode: JSON modifications to apply")
    parser.add_argument("--template", help="Base template path (optional)")

    args = parser.parse_args()

    try:
        if args.clone:
            print(f"Cloning workbook from: {args.clone}")
            wb = clone_workbook(args.clone, args.json)
        else:
            if not args.input:
                print("Error: Input JSON file required (or use --clone)")
                sys.exit(1)

            input_path = Path(args.input)
            if not input_path.exists():
                print(f"Error: Input file not found: {args.input}")
                sys.exit(1)

            print(f"Generating workbook from: {args.input}")
            wb = generate_workbook(str(input_path), args.template)

        # Save workbook
        output_path = Path(args.output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        print(f"\nCreated: {args.output}")
        print(f"Sheets: {len(wb.sheetnames)}")
        for name in wb.sheetnames:
            print(f"  - {name}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
